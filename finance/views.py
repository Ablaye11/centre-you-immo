from django.views.generic import TemplateView, CreateView, View, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.utils import timezone
from .models import Invoice, Expense, Payment
from tenants.models import Tenant, Shop
from django.db.models import Sum, Prefetch, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class FinanceOverviewView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/finance_overview.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'finance'

        active_mall = self.request.active_mall

        invoices_qs = Invoice.objects.select_related('tenant', 'shop')
        expenses_qs = Expense.objects.all()

        if active_mall:
            invoices_qs = invoices_qs.filter(shop__mall=active_mall)
            expenses_qs = expenses_qs.filter(mall=active_mall)

        # Totals
        context['total_invoiced'] = invoices_qs.aggregate(total=Sum('amount'))['total'] or 0
        context['total_collected'] = invoices_qs.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
        context['total_pending'] = invoices_qs.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
        context['total_overdue'] = invoices_qs.filter(status='overdue').aggregate(total=Sum('amount'))['total'] or 0
        context['total_expenses'] = expenses_qs.aggregate(total=Sum('amount'))['total'] or 0

        # Balance
        context['net_balance'] = context['total_collected'] - context['total_expenses']

        # Search filter
        q = self.request.GET.get('q')
        invoices = invoices_qs.order_by('-issue_date')
        expenses = expenses_qs.order_by('-date')

        if q:
            invoices = invoices.filter(
                Q(invoice_number__icontains=q) |
                Q(tenant__first_name__icontains=q) |
                Q(tenant__last_name__icontains=q) |
                Q(shop__shop_number__icontains=q) |
                Q(shop__name__icontains=q)
            )
            expenses = expenses.filter(
                Q(title__icontains=q) |
                Q(category__icontains=q) |
                Q(supplier__icontains=q)
            )

        context['invoices'] = invoices[:20]
        context['expenses'] = expenses[:20]
        return context


class InvoiceCreateView(LoginRequiredMixin, CreateView):
    model = Invoice
    template_name = 'finance/invoice_form.html'
    fields = ['invoice_number', 'tenant', 'shop', 'invoice_type', 'amount', 'issue_date', 'due_date', 'status', 'description']
    success_url = reverse_lazy('finance_overview')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'finance'
        context['title'] = "Emettre une Facture"
        return context

    def form_valid(self, form):
        messages.success(self.request, "La facture a ete emise avec succes.")
        return super().form_valid(form)


class ExpenseCreateView(LoginRequiredMixin, CreateView):
    model = Expense
    template_name = 'finance/expense_form.html'
    fields = ['title', 'category', 'amount', 'date', 'supplier', 'description']
    success_url = reverse_lazy('finance_overview')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'finance'
        context['title'] = "Enregistrer une Depense"
        return context

    def form_valid(self, form):
        if self.request.active_mall:
            form.instance.mall = self.request.active_mall
        messages.success(self.request, "La depense a ete enregistree avec succes.")
        return super().form_valid(form)


class MarkInvoicePaidView(LoginRequiredMixin, View):
    """Mark an invoice as paid with current date and create a Payment record."""

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        if invoice.status != 'paid':
            invoice.status = 'paid'
            invoice.paid_date = timezone.now().date()
            invoice.save()
            # Create corresponding payment record
            from finance.models import Payment
            Payment.objects.create(
                invoice=invoice,
                amount=invoice.amount,
                payment_date=invoice.paid_date,
                method=request.POST.get('payment_method', 'cash'),
                notes=request.POST.get('notes', ''),
            )
            
            # Notify admins, managers and accountants
            from django.contrib.auth.models import User
            from core.models import Notification
            
            admins = User.objects.filter(Q(is_superuser=True) | Q(profile__role__in=['admin', 'manager', 'accountant']))
            for admin in admins:
                if admin != request.user:
                    Notification.objects.create(
                        user=admin,
                        title="Paiement Reçu",
                        message=f"Paiement de {invoice.amount:,} FCFA enregistré pour la facture {invoice.invoice_number} ({invoice.tenant.full_name})",
                        notif_type='success'
                    )
            
            messages.success(request, f"✅ Facture {invoice.invoice_number} payée — {invoice.amount} FCFA enregistrés.")
        else:
            messages.info(request, f"Facture {invoice.invoice_number} est deja payee.")
        return redirect('finance_overview')


class GenerateMonthlyInvoicesView(LoginRequiredMixin, View):
    """Auto-generate monthly rent invoices for all active leases of the active mall."""

    def post(self, request):
        from tenants.models import Lease
        import datetime

        active_mall = request.active_mall
        if not active_mall:
            messages.error(request, "Aucun centre commercial actif sélectionné.")
            return redirect('finance_overview')

        today = timezone.now().date()
        month_label = today.strftime('%B %Y')
        first_day = today.replace(day=1)
        # Due date = last day of month
        if today.month == 12:
            last_day = today.replace(day=31)
        else:
            last_day = today.replace(month=today.month + 1, day=1) - datetime.timedelta(days=1)

        active_leases = Lease.objects.filter(shop__mall=active_mall, status='active').select_related('tenant', 'shop')
        created_count = 0
        skipped_count = 0

        for lease in active_leases:
            prefix = f"FAC-{today.year}-"
            already_exists = Invoice.objects.filter(
                tenant=lease.tenant,
                shop=lease.shop,
                invoice_type='rent',
                issue_date__year=today.year,
                issue_date__month=today.month,
            ).exists()

            if already_exists:
                skipped_count += 1
                continue

            last = Invoice.objects.filter(invoice_number__startswith=prefix).order_by('-invoice_number').first()
            seq = 1
            if last:
                try:
                    seq = int(last.invoice_number.split('-')[-1]) + 1
                except ValueError:
                    pass
            inv_number = f"{prefix}{seq:04d}"

            Invoice.objects.create(
                invoice_number=inv_number,
                tenant=lease.tenant,
                shop=lease.shop,
                invoice_type='rent',
                amount=lease.monthly_rent,
                issue_date=first_day,
                due_date=last_day,
                status='pending',
                description=f"Loyer mensuel — {month_label} — {lease.shop.name}",
            )
            created_count += 1

        if created_count > 0:
            messages.success(request, f"✅ {created_count} facture(s) de loyer générées pour {month_label}.")
        if skipped_count > 0:
            messages.info(request, f"ℹ️ {skipped_count} bail(s) avaient déjà une facture ce mois-ci.")
        if created_count == 0 and skipped_count == 0:
            messages.warning(request, "Aucun bail actif trouvé pour ce centre.")

        return redirect('finance_overview')


class ExportInvoicesExcelView(LoginRequiredMixin, View):
    """Export all invoices to an Excel file."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Factures YOU IMMO"

        # Header style
        header_fill = PatternFill(start_color="D2691E", end_color="D2691E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        headers = ["N° Facture", "Locataire", "Boutique", "Type", "Montant (FCFA)", "Date Emission", "Date Echeance", "Date Paiement", "Statut"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Column widths
        col_widths = [15, 25, 20, 15, 18, 16, 16, 16, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Data rows with select_related to avoid N+1 queries
        status_map = {'paid': 'Payee', 'pending': 'En attente', 'overdue': 'En retard', 'cancelled': 'Annulee'}
        for inv in Invoice.objects.select_related('tenant', 'shop').order_by('-issue_date'):
            ws.append([
                inv.invoice_number,
                inv.tenant.full_name,
                f"{inv.shop.shop_number} - {inv.shop.name}",
                inv.get_invoice_type_display(),
                float(inv.amount),
                str(inv.issue_date),
                str(inv.due_date),
                str(inv.paid_date) if inv.paid_date else "-",
                status_map.get(inv.status, inv.status),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="factures_youimmo.xlsx"'
        wb.save(response)
        return response


class ExportTenantsExcelView(LoginRequiredMixin, View):
    """Export all tenants to an Excel file."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Locataires YOU IMMO"

        header_fill = PatternFill(start_color="D2691E", end_color="D2691E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        headers = ["Nom Complet", "Telephone", "Email", "Boutique", "Loyer Mensuel (FCFA)", "Depot Garantie (FCFA)", "Date Debut Bail", "Date Fin Bail", "Statut"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        col_widths = [25, 16, 28, 22, 22, 22, 16, 16, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        from tenants.models import Lease
        active_leases_prefetch = Prefetch(
            'leases',
            queryset=Lease.objects.select_related('shop').filter(status='active'),
            to_attr='active_leases'
        )
        tenants = Tenant.objects.prefetch_related(active_leases_prefetch).order_by('last_name')

        for tenant in tenants:
            lease = tenant.active_leases[0] if tenant.active_leases else None
            ws.append([
                tenant.full_name,
                tenant.phone,
                tenant.email or "-",
                f"{lease.shop.shop_number} - {lease.shop.name}" if lease else "-",
                float(lease.monthly_rent) if lease else "-",
                float(lease.deposit) if lease else "-",
                str(lease.start_date) if lease else "-",
                str(lease.end_date) if lease else "-",
                "Actif" if lease else "Inactif",
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="locataires_youimmo.xlsx"'
        wb.save(response)
        return response


class InvoiceDeleteView(LoginRequiredMixin, DeleteView):
    model = Invoice
    template_name = 'finance/invoice_confirm_delete.html'
    success_url = reverse_lazy('finance_overview')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'finance'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"La facture '{self.object.invoice_number}' a été supprimée.")
        return super().form_valid(form)


class ExpenseDeleteView(LoginRequiredMixin, DeleteView):
    model = Expense
    template_name = 'finance/expense_confirm_delete.html'
    success_url = reverse_lazy('finance_overview')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'finance'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"La dépense '{self.object.title}' a été supprimée.")
        return super().form_valid(form)


class ReportOverviewView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/report_overview.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'reports'

        active_mall = self.request.active_mall
        if not active_mall:
            context['no_active_mall'] = True
            return context

        period = self.request.GET.get('period', 'month')
        today = timezone.now().date()

        if period == 'week':
            start_date = today - timezone.timedelta(days=today.weekday())
            end_date = start_date + timezone.timedelta(days=6)
            period_label = "Cette semaine"
        elif period == 'last_month':
            first_day_of_this_month = today.replace(day=1)
            end_date = first_day_of_this_month - timezone.timedelta(days=1)
            start_date = end_date.replace(day=1)
            period_label = end_date.strftime('%B %Y')
        else:  # month
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(day=31)
            else:
                end_date = (today.replace(month=today.month + 1, day=1) - timezone.timedelta(days=1))
            period_label = today.strftime('%B %Y')

        # Basic counts
        total_shops = Shop.objects.filter(mall=active_mall).count()
        occupied_shops = Shop.objects.filter(mall=active_mall, status='occupied').count()
        occupancy_rate = int((occupied_shops / total_shops * 100)) if total_shops > 0 else 0

        # Financials in period
        invoices_in_period = Invoice.objects.filter(shop__mall=active_mall, issue_date__range=(start_date, end_date))
        expenses_in_period = Expense.objects.filter(mall=active_mall, date__range=(start_date, end_date))

        revenue_collected = Payment.objects.filter(
            invoice__shop__mall=active_mall,
            payment_date__range=(start_date, end_date)
        ).aggregate(total=Sum('amount'))['total'] or 0

        expenses_total = expenses_in_period.aggregate(total=Sum('amount'))['total'] or 0
        net_income = revenue_collected - expenses_total

        # Invoices stats
        invoiced_amount = invoices_in_period.aggregate(total=Sum('amount'))['total'] or 0
        pending_amount = invoices_in_period.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
        overdue_amount = invoices_in_period.filter(status='overdue').aggregate(total=Sum('amount'))['total'] or 0
        paid_amount = invoices_in_period.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0

        # Maintenance
        from maintenance.models import MaintenanceRequest
        maintenance_qs = MaintenanceRequest.objects.filter(mall=active_mall, created_at__date__range=(start_date, end_date))
        maintenance_count = maintenance_qs.count()
        maintenance_actual_cost = maintenance_qs.aggregate(total=Sum('actual_cost'))['total'] or 0

        # Top paying tenants / Leases in active mall
        from tenants.models import Lease
        active_leases = Lease.objects.filter(shop__mall=active_mall, status='active').select_related('tenant', 'shop').order_by('-monthly_rent')[:5]

        # Recent activities
        recent_payments = Payment.objects.filter(
            invoice__shop__mall=active_mall,
            payment_date__range=(start_date, end_date)
        ).select_related('invoice__tenant', 'invoice__shop').order_by('-payment_date')[:5]

        # Salaries — employees of this mall and their salary costs
        from employees.models import Employee
        active_employees = Employee.objects.filter(mall=active_mall, status='active')
        employees_count = active_employees.count()
        salaries_total = active_employees.aggregate(total=Sum('salary'))['total'] or 0

        # Salary expenses already registered in the period
        salary_expenses_in_period = expenses_in_period.filter(category='salary')
        salary_expenses_total = salary_expenses_in_period.aggregate(total=Sum('amount'))['total'] or 0

        # Expense breakdown by category for inventory
        from django.db.models import Sum as DSum
        expense_by_category = (
            expenses_in_period
            .values('category')
            .annotate(total=DSum('amount'))
            .order_by('-total')
        )

        context.update({
            'period': period,
            'period_label': period_label,
            'start_date': start_date,
            'end_date': end_date,
            'total_shops': total_shops,
            'occupied_shops': occupied_shops,
            'occupancy_rate': occupancy_rate,
            'revenue_collected': revenue_collected,
            'expenses_total': expenses_total,
            'net_income': net_income,
            'invoiced_amount': invoiced_amount,
            'pending_amount': pending_amount,
            'overdue_amount': overdue_amount,
            'paid_amount': paid_amount,
            'maintenance_count': maintenance_count,
            'maintenance_actual_cost': maintenance_actual_cost,
            'active_leases': active_leases,
            'recent_payments': recent_payments,
            'invoices': invoices_in_period.order_by('-issue_date')[:10],
            'expenses': expenses_in_period.order_by('-date')[:10],
            # Salary data
            'employees_count': employees_count,
            'salaries_total': salaries_total,
            'salary_expenses_total': salary_expenses_total,
            'active_employees': active_employees,
            'expense_by_category': expense_by_category,
        })
        return context


class ExportReportExcelView(LoginRequiredMixin, View):
    """Export activity reports for the active mall to an Excel workbook with multiple sheets."""

    def get(self, request):
        active_mall = request.active_mall
        if not active_mall:
            messages.error(request, "Aucun centre commercial actif sélectionné.")
            return redirect('finance_overview')

        period = request.GET.get('period', 'month')
        today = timezone.now().date()

        if period == 'week':
            start_date = today - timezone.timedelta(days=today.weekday())
            end_date = start_date + timezone.timedelta(days=6)
            period_label = f"Semaine du {start_date} au {end_date}"
        elif period == 'last_month':
            first_day_of_this_month = today.replace(day=1)
            end_date = first_day_of_this_month - timezone.timedelta(days=1)
            start_date = end_date.replace(day=1)
            period_label = end_date.strftime('%B %Y')
        else:  # month
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(day=31)
            else:
                end_date = (today.replace(month=today.month + 1, day=1) - timezone.timedelta(days=1))
            period_label = today.strftime('%B %Y')

        wb = openpyxl.Workbook()

        # Styles
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Blue
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        title_font = Font(bold=True, size=14, color="1E3A8A")
        bold_font = Font(bold=True)

        # SHEET 1: RESUME
        ws1 = wb.active
        ws1.title = "Résumé"
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 25

        ws1.append([])
        ws1.append([f"RAPPORT D'ACTIVITÉ - {active_mall.name.upper()}"])
        ws1.cell(row=2, column=1).font = title_font
        ws1.append([f"Période : {period_label} ({start_date} au {end_date})"])
        ws1.append([])

        # Gather stats
        total_shops = Shop.objects.filter(mall=active_mall).count()
        occupied_shops = Shop.objects.filter(mall=active_mall, status='occupied').count()
        occupancy_rate = f"{(occupied_shops / total_shops * 100):.1f}%" if total_shops > 0 else "0%"

        revenue_collected = Payment.objects.filter(
            invoice__shop__mall=active_mall,
            payment_date__range=(start_date, end_date)
        ).aggregate(total=Sum('amount'))['total'] or 0

        expenses_total = Expense.objects.filter(
            mall=active_mall,
            date__range=(start_date, end_date)
        ).aggregate(total=Sum('amount'))['total'] or 0

        net_income = revenue_collected - expenses_total

        from maintenance.models import MaintenanceRequest
        maintenance_count = MaintenanceRequest.objects.filter(
            mall=active_mall,
            created_at__date__range=(start_date, end_date)
        ).count()

        ws1.append(["Indicateur", "Valeur"])
        for col in [1, 2]:
            cell = ws1.cell(row=ws1.max_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws1.append(["Boutiques Totales", total_shops])
        ws1.append(["Boutiques Occupées", occupied_shops])
        ws1.append(["Taux d'Occupation", occupancy_rate])
        ws1.append(["Recettes (Loyers Encaissés)", float(revenue_collected)])
        ws1.cell(row=ws1.max_row, column=2).number_format = '#,##0'
        ws1.append(["Dépenses Opérationnelles", float(expenses_total)])
        ws1.cell(row=ws1.max_row, column=2).number_format = '#,##0'
        ws1.append(["Résultat Net", float(net_income)])
        ws1.cell(row=ws1.max_row, column=2).number_format = '#,##0'
        ws1.cell(row=ws1.max_row, column=2).font = bold_font
        ws1.append(["Incidents Maintenance", maintenance_count])

        # SHEET 2: FACTURES
        ws2 = wb.create_sheet(title="Factures")
        headers2 = ["N° Facture", "Locataire", "Boutique", "Type", "Montant (FCFA)", "Date Émission", "Date Échéance", "Statut"]
        ws2.append(headers2)
        for col_num, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 15
        ws2.column_dimensions['H'].width = 12

        status_map = {'paid': 'Payée', 'pending': 'En attente', 'overdue': 'En retard', 'cancelled': 'Annulée'}
        invoices = Invoice.objects.filter(shop__mall=active_mall, issue_date__range=(start_date, end_date)).select_related('tenant', 'shop').order_by('-issue_date')
        for inv in invoices:
            ws2.append([
                inv.invoice_number,
                inv.tenant.full_name,
                f"{inv.shop.shop_number} - {inv.shop.name}",
                inv.get_invoice_type_display(),
                float(inv.amount),
                str(inv.issue_date),
                str(inv.due_date),
                status_map.get(inv.status, inv.status)
            ])
            ws2.cell(row=ws2.max_row, column=5).number_format = '#,##0'

        # SHEET 3: DEPENSES
        ws3 = wb.create_sheet(title="Dépenses")
        headers3 = ["Titre", "Catégorie", "Montant (FCFA)", "Date", "Fournisseur", "Description"]
        ws3.append(headers3)
        for col_num, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 18
        ws3.column_dimensions['D'].width = 15
        ws3.column_dimensions['E'].width = 25
        ws3.column_dimensions['F'].width = 35

        expenses = Expense.objects.filter(mall=active_mall, date__range=(start_date, end_date)).order_by('-date')
        for exp in expenses:
            ws3.append([
                exp.title,
                exp.get_category_display(),
                float(exp.amount),
                str(exp.date),
                exp.supplier or "-",
                exp.description or ""
            ])
            ws3.cell(row=ws3.max_row, column=3).number_format = '#,##0'

        # SHEET 4: MAINTENANCE
        ws4 = wb.create_sheet(title="Maintenance")
        headers4 = ["Titre", "Zone", "Priorité", "Statut", "Assigné à", "Coût Réel (FCFA)", "Date Signalement"]
        ws4.append(headers4)
        for col_num, h in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 12
        ws4.column_dimensions['D'].width = 12
        ws4.column_dimensions['E'].width = 20
        ws4.column_dimensions['F'].width = 18
        ws4.column_dimensions['G'].width = 18

        maintenance_requests = MaintenanceRequest.objects.filter(mall=active_mall, created_at__date__range=(start_date, end_date)).order_by('-created_at')
        for req in maintenance_requests:
            ws4.append([
                req.title,
                req.get_zone_display(),
                req.get_priority_display(),
                req.get_status_display(),
                req.assigned_to or "-",
                float(req.actual_cost) if req.actual_cost else 0,
                req.created_at.strftime('%Y-%m-%d %H:%M')
            ])
            ws4.cell(row=ws4.max_row, column=6).number_format = '#,##0'

        # SHEET 5: BAUX ACTIFS
        ws5 = wb.create_sheet(title="Baux Actifs")
        headers5 = ["Locataire", "Boutique", "Loyer Mensuel (FCFA)", "Caution (FCFA)", "Date Début", "Date Fin"]
        ws5.append(headers5)
        for col_num, h in enumerate(headers5, 1):
            cell = ws5.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws5.column_dimensions['A'].width = 25
        ws5.column_dimensions['B'].width = 20
        ws5.column_dimensions['C'].width = 18
        ws5.column_dimensions['D'].width = 18
        ws5.column_dimensions['E'].width = 15
        ws5.column_dimensions['F'].width = 15

        from tenants.models import Lease
        leases = Lease.objects.filter(shop__mall=active_mall, status='active').select_related('tenant', 'shop').order_by('shop__shop_number')
        for lease in leases:
            ws5.append([
                lease.tenant.full_name,
                f"{lease.shop.shop_number} - {lease.shop.name}",
                float(lease.monthly_rent),
                float(lease.deposit),
                str(lease.start_date),
                str(lease.end_date)
            ])
            ws5.cell(row=ws5.max_row, column=3).number_format = '#,##0'
            ws5.cell(row=ws5.max_row, column=4).number_format = '#,##0'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{active_mall.name.replace(" ", "_")}_{period}.xlsx"'
        wb.save(response)
        return response


class DistributeChargesView(LoginRequiredMixin, View):
    def get(self, request):
        active_mall = request.active_mall
        if not active_mall:
            messages.error(request, "Veuillez sélectionner un centre commercial.")
            return redirect('finance_overview')
        from tenants.models import Lease
        active_leases = Lease.objects.filter(shop__mall=active_mall, status='active').select_related('tenant', 'shop').order_by('shop__shop_number')
        categories = [
            ('energy', 'Énergie & Eau'),
            ('cleaning', 'Nettoyage'),
            ('security', 'Sécurité'),
            ('maintenance', 'Maintenance'),
            ('insurance', 'Assurance'),
            ('marketing', 'Marketing'),
            ('other', 'Autre'),
        ]
        total_surface = sum(float(l.shop.surface or 0) for l in active_leases if l.shop.surface)
        context = {
            'active_menu': 'finance',
            'active_mall': active_mall,
            'active_leases': active_leases,
            'categories': categories,
            'total_surface': total_surface
        }
        return render(request, 'finance/distribute_charges.html', context)

    def post(self, request):
        from tenants.models import Lease
        active_mall = request.active_mall
        if not active_mall:
            messages.error(request, "Veuillez sélectionner un centre commercial.")
            return redirect('finance_overview')
        
        total_amount_raw = request.POST.get('total_amount', '').replace(' ', '').replace(',', '')
        category = request.POST.get('category', 'other')
        description = request.POST.get('description', '')
        issue_date_str = request.POST.get('issue_date')
        due_date_str = request.POST.get('due_date')
        
        try:
            total_amount = int(total_amount_raw)
        except (ValueError, TypeError):
            messages.error(request, "Montant invalide.")
            return redirect('distribute_charges')
            
        issue_date = timezone.datetime.strptime(issue_date_str, '%Y-%m-%d').date() if issue_date_str else timezone.now().date()
        due_date = timezone.datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else timezone.now().date().replace(day=28)
        
        active_leases = Lease.objects.filter(shop__mall=active_mall, status='active').select_related('tenant', 'shop')
        total_surface = sum(float(l.shop.surface) for l in active_leases if l.shop.surface)
        
        if total_surface <= 0:
            messages.error(request, "Aucun bail actif avec surface définie.")
            return redirect('distribute_charges')
            
        cost_per_m2 = total_amount / total_surface
        cat_labels = {
            'energy': 'Énergie & Eau',
            'cleaning': 'Nettoyage',
            'security': 'Sécurité',
            'maintenance': 'Maintenance',
            'insurance': 'Assurance',
            'marketing': 'Marketing',
            'other': 'Autre'
        }
        cat_label = cat_labels.get(category, 'Charges')
        count = 0
        
        for lease in active_leases:
            surface = float(lease.shop.surface or 0)
            if surface <= 0:
                continue
            share = round(surface * cost_per_m2)
            if share <= 0:
                continue
            
            # Generate invoice number
            today = timezone.now()
            prefix = f"FAC-{today.year}-"
            last = Invoice.objects.filter(invoice_number__startswith=prefix).order_by('-invoice_number').first()
            seq = int(last.invoice_number.split('-')[-1]) + 1 if last else 1
            invoice_number = f"{prefix}{seq:04d}"
            
            Invoice.objects.create(
                invoice_number=invoice_number,
                tenant=lease.tenant,
                shop=lease.shop,
                invoice_type='charges',
                amount=share,
                issue_date=issue_date,
                due_date=due_date,
                status='pending',
                description=description or f"{cat_label} — Répartition prorata ({surface}m² / {total_surface:.1f}m² total)",
            )
            count += 1
            
        messages.success(request, f"✅ {count} facture(s) de charges générée(s) — {total_amount:,} FCFA répartis entre les locataires actifs de {active_mall.name}.")
        return redirect('finance_overview')


class GenerateMonthlySalariesView(LoginRequiredMixin, View):
    """Create one Expense(salary) per active employee of the active mall for the current month."""

    def post(self, request):
        from employees.models import Employee
        active_mall = request.active_mall
        if not active_mall:
            messages.error(request, "Veuillez sélectionner un centre commercial.")
            return redirect('finance_overview')

        today = timezone.now().date()
        month_label = today.strftime('%B %Y')
        first_day = today.replace(day=1)

        active_employees = Employee.objects.filter(mall=active_mall, status='active')
        if not active_employees.exists():
            messages.warning(request, "Aucun employé actif trouvé pour ce centre commercial.")
            return redirect('finance_overview')

        created = 0
        skipped = 0
        for emp in active_employees:
            title = f"Salaire — {emp.full_name} ({month_label})"
            # Avoid duplicate: check if salary expense already exists this month for this employee
            already = Expense.objects.filter(
                mall=active_mall,
                category='salary',
                title=title,
                date__year=today.year,
                date__month=today.month,
            ).exists()
            if already:
                skipped += 1
                continue
            Expense.objects.create(
                mall=active_mall,
                title=title,
                category='salary',
                amount=emp.salary,
                date=first_day,
                description=f"Salaire mensuel — Poste : {emp.position} | Contrat : {emp.get_contract_type_display()}",
                supplier=emp.full_name,
            )
            created += 1

        if created:
            messages.success(request, f"✅ {created} fiche(s) de salaire générée(s) pour {month_label} ({active_mall.name}).")
        if skipped:
            messages.info(request, f"ℹ️ {skipped} employé(s) déjà traité(s) ce mois-ci — non dupliqués.")
        return redirect('finance_overview')
