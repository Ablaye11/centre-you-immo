from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.db import transaction
from django.shortcuts import redirect, render
from django.utils import timezone
from .models import Tenant, Shop, Lease, Floor, Mall
from .forms import ShopForm, TenantForm, FloorForm, MallForm
from finance.models import Invoice, Payment


def _generate_invoice_number():
    """Generate a unique sequential invoice number like FAC-2024-0042."""
    today = timezone.now()
    prefix = f"FAC-{today.year}-"
    invoice_numbers = Invoice.objects.filter(invoice_number__startswith=prefix).values_list('invoice_number', flat=True)
    
    max_seq = 0
    for num in invoice_numbers:
        try:
            seq = int(num.split('-')[-1])
            if seq > max_seq:
                max_seq = seq
        except ValueError:
            pass
            
    return f"{prefix}{(max_seq + 1):04d}"


class TenantListView(LoginRequiredMixin, ListView):
    model = Tenant
    template_name = 'tenants/tenant_list.html'
    context_object_name = 'tenants'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        active_mall = self.request.active_mall
        context['shops'] = Shop.objects.filter(mall=active_mall) if active_mall else Shop.objects.all()
        return context

    def get_queryset(self):
        active_mall = self.request.active_mall
        # Tenants that have at least one lease in the active mall
        if active_mall:
            queryset = Tenant.objects.filter(leases__shop__mall=active_mall).distinct().prefetch_related('leases')
        else:
            queryset = Tenant.objects.prefetch_related('leases')
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(first_name__icontains=query) | queryset.filter(last_name__icontains=query) | queryset.filter(company_name__icontains=query)
        return queryset


class TenantDetailView(LoginRequiredMixin, DetailView):
    model = Tenant
    template_name = 'tenants/tenant_detail.html'
    context_object_name = 'tenant'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        context['leases'] = self.object.leases.all()
        return context


class TenantCreateView(LoginRequiredMixin, CreateView):
    model = Tenant
    template_name = 'tenants/tenant_form.html'
    form_class = TenantForm
    success_url = reverse_lazy('tenant_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['active_mall'] = self.request.active_mall
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        context['title'] = "Nouveau locataire"
        context['is_new'] = True
        return context

    @transaction.atomic
    def form_valid(self, form):
        tenant = form.save()
        shop = form.cleaned_data.get('shop')
        monthly_rent = form.cleaned_data.get('monthly_rent')
        deposit = form.cleaned_data.get('deposit')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        initial_payment = form.cleaned_data.get('initial_payment', 'none')
        payment_method = form.cleaned_data.get('payment_method', 'cash')

        lease = None
        if shop and monthly_rent and start_date and end_date:
            lease = Lease.objects.create(
                shop=shop,
                tenant=tenant,
                start_date=start_date,
                end_date=end_date,
                monthly_rent=monthly_rent,
                deposit=deposit or 0,
                status='active',
            )
            # Mark shop as occupied
            shop.status = 'occupied'
            shop.save(update_fields=['status'])

            today = timezone.now().date()

            def _make_invoice(inv_type, amount, description, paid=False):
                inv = Invoice.objects.create(
                    invoice_number=_generate_invoice_number(),
                    tenant=tenant,
                    shop=shop,
                    invoice_type=inv_type,
                    amount=amount,
                    issue_date=today,
                    due_date=today,
                    status='paid' if paid else 'pending',
                    paid_date=today if paid else None,
                    description=description,
                )
                if paid:
                    Payment.objects.create(
                        invoice=inv,
                        amount=amount,
                        payment_date=today,
                        method=payment_method,
                        notes=f"Paiement initial à la signature du bail",
                    )
                return inv

            # Generate invoices based on initial_payment choice
            if initial_payment in ('deposit', 'both') and deposit:
                _make_invoice('charges', deposit, f"Caution - Bail {shop}", paid=True)
            elif deposit:
                _make_invoice('charges', deposit, f"Caution - Bail {shop}", paid=False)

            if initial_payment in ('rent', 'both') and monthly_rent:
                _make_invoice('rent', monthly_rent, f"Premier loyer - {start_date.strftime('%B %Y')}", paid=True)
            else:
                _make_invoice('rent', monthly_rent, f"Loyer - {start_date.strftime('%B %Y')}", paid=False)

            messages.success(self.request, f"✅ Locataire '{tenant.full_name}' créé, bail signé pour {shop.name}, factures générées.")
        else:
            messages.success(self.request, f"✅ Locataire '{tenant.full_name}' créé avec succès.")

        return redirect(self.success_url)


class TenantUpdateView(LoginRequiredMixin, UpdateView):
    model = Tenant
    template_name = 'tenants/tenant_form.html'
    form_class = TenantForm
    success_url = reverse_lazy('tenant_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        context['title'] = "Modifier le locataire"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Les informations du locataire ont ete mises a jour.")
        return super().form_valid(form)


class TenantDeleteView(LoginRequiredMixin, DeleteView):
    model = Tenant
    template_name = 'tenants/tenant_confirm_delete.html'
    success_url = reverse_lazy('tenant_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Le locataire '{self.object.full_name}' a été supprimé.")
        return super().form_valid(form)


class ShopListView(LoginRequiredMixin, TemplateView):
    """Visual map of all shops showing occupied vs available."""
    template_name = 'tenants/shop_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'

        active_mall = self.request.active_mall
        if active_mall:
            all_shops = Shop.objects.filter(mall=active_mall).select_related('floor').order_by('floor', 'shop_number')
        else:
            all_shops = Shop.objects.all().select_related('floor').order_by('floor', 'shop_number')
        context['shops'] = all_shops

        # Stats
        total = all_shops.count()
        occupied = all_shops.filter(status='occupied').count()
        available = all_shops.filter(status='available').count()
        maintenance = all_shops.filter(status='maintenance').count()

        context['total_shops'] = total
        context['occupied_shops'] = occupied
        context['available_shops'] = available
        context['maintenance_shops'] = maintenance
        context['occupancy_rate'] = round((occupied / total * 100)) if total > 0 else 0

        # Group by floor
        floors = {}
        for shop in all_shops:
            f = shop.floor.name if shop.floor else 'Non assigné'
            if f not in floors:
                floors[f] = []
            floors[f].append(shop)
        context['floors'] = floors

        return context




class ShopCreateView(LoginRequiredMixin, CreateView):
    """Create a new shop/boutique."""
    model = Shop
    template_name = 'tenants/shop_form.html'
    form_class = ShopForm
    success_url = reverse_lazy('shop_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['active_mall'] = self.request.active_mall
        return kwargs

    def form_valid(self, form):
        if self.request.active_mall:
            form.instance.mall = self.request.active_mall
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'shops'
        context['title'] = 'Ajouter une Boutique'
        context['is_new'] = True
        return context

    def form_valid(self, form):
        messages.success(self.request, f"La boutique '{form.instance.name}' a ete ajoutee avec succes.")
        return super().form_valid(form)


class ShopUpdateView(LoginRequiredMixin, UpdateView):
    """Edit an existing shop."""
    model = Shop
    template_name = 'tenants/shop_form.html'
    form_class = ShopForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['active_mall'] = self.request.active_mall
        return kwargs
    success_url = reverse_lazy('shop_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'shops'
        context['title'] = f'Modifier - {self.object.name}'
        context['is_new'] = False
        return context

    def form_valid(self, form):
        messages.success(self.request, f"La boutique '{form.instance.name}' a ete mise a jour.")
        return super().form_valid(form)


class ShopDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a shop."""
    model = Shop
    template_name = 'tenants/shop_confirm_delete.html'
    success_url = reverse_lazy('shop_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'shops'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"La boutique '{self.object.name}' a ete supprimee.")
        return super().form_valid(form)


# ---- Floor Views ----

class FloorListView(LoginRequiredMixin, ListView):
    model = Floor
    template_name = 'tenants/floor_list.html'
    context_object_name = 'floors'

    def get_queryset(self):
        active_mall = self.request.active_mall
        if active_mall:
            return Floor.objects.filter(mall=active_mall)
        return Floor.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'floors'
        return context

class FloorCreateView(LoginRequiredMixin, CreateView):
    model = Floor
    form_class = FloorForm
    template_name = 'tenants/floor_form.html'
    success_url = reverse_lazy('floor_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'floors'
        context['title'] = "Ajouter un Étage"
        return context

    def form_valid(self, form):
        if self.request.active_mall:
            form.instance.mall = self.request.active_mall
        messages.success(self.request, f"L'étage '{form.instance.name}' a été ajouté.")
        return super().form_valid(form)

class FloorUpdateView(LoginRequiredMixin, UpdateView):
    model = Floor
    form_class = FloorForm
    template_name = 'tenants/floor_form.html'
    success_url = reverse_lazy('floor_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'floors'
        context['title'] = f"Modifier - {self.object.name}"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"L'étage '{form.instance.name}' a été mis à jour.")
        return super().form_valid(form)

class FloorDeleteView(LoginRequiredMixin, DeleteView):
    model = Floor
    template_name = 'tenants/floor_confirm_delete.html'
    success_url = reverse_lazy('floor_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'floors'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"L'étage '{self.object.name}' a été supprimé.")
        return super().form_valid(form)


# ---- Mall (Centre Commercial) CRUD Views ----

class MallListView(LoginRequiredMixin, ListView):
    model = Mall
    template_name = 'tenants/mall_list.html'
    context_object_name = 'malls'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'malls'
        return context


class MallCreateView(LoginRequiredMixin, CreateView):
    model = Mall
    form_class = MallForm
    template_name = 'tenants/mall_form.html'
    success_url = reverse_lazy('mall_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'malls'
        context['title'] = "Ajouter un Centre Commercial"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"✅ Centre commercial '{form.instance.name}' créé avec succès.")
        return super().form_valid(form)


class MallUpdateView(LoginRequiredMixin, UpdateView):
    model = Mall
    form_class = MallForm
    template_name = 'tenants/mall_form.html'
    success_url = reverse_lazy('mall_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'malls'
        context['title'] = f"Modifier - {self.object.name}"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Centre commercial '{form.instance.name}' mis à jour.")
        return super().form_valid(form)


class MallDeleteView(LoginRequiredMixin, DeleteView):
    model = Mall
    template_name = 'tenants/mall_confirm_delete.html'
    success_url = reverse_lazy('mall_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'malls'
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Centre commercial '{self.object.name}' supprimé.")
        return super().form_valid(form)


class LeasePrintView(LoginRequiredMixin, DetailView):
    model = Lease
    template_name = 'tenants/lease_print.html'
    context_object_name = 'lease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_menu'] = 'tenants'
        context['today'] = timezone.now().date()
        return context


class AddLeaseView(LoginRequiredMixin, View):
    """Assign a new shop (bail) to an existing tenant."""

    def get(self, request, pk):
        tenant = Tenant.objects.get(pk=pk)
        active_mall = request.active_mall
        if active_mall:
            available_shops = Shop.objects.filter(mall=active_mall, status='available')
        else:
            available_shops = Shop.objects.filter(status='available')
        return render(request, 'tenants/add_lease.html', {
            'tenant': tenant,
            'available_shops': available_shops,
            'active_menu': 'tenants',
            'today': timezone.now().date().isoformat(),
        })

    @transaction.atomic
    def post(self, request, pk):
        tenant = Tenant.objects.get(pk=pk)
        active_mall = request.active_mall

        shop_id = request.POST.get('shop')
        monthly_rent = request.POST.get('monthly_rent')
        deposit = request.POST.get('deposit') or 0
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        initial_payment = request.POST.get('initial_payment', 'none')
        payment_method = request.POST.get('payment_method', 'cash')

        if not (shop_id and monthly_rent and start_date and end_date):
            messages.error(request, "Veuillez remplir tous les champs obligatoires : boutique, loyer, dates de début et de fin.")
            if active_mall:
                available_shops = Shop.objects.filter(mall=active_mall, status='available')
            else:
                available_shops = Shop.objects.filter(status='available')
            return render(request, 'tenants/add_lease.html', {
                'tenant': tenant,
                'available_shops': available_shops,
                'active_menu': 'tenants',
                'today': timezone.now().date().isoformat(),
            })

        try:
            shop = Shop.objects.get(pk=shop_id)
        except Shop.DoesNotExist:
            messages.error(request, "Boutique introuvable.")
            return redirect('tenant_detail', pk=pk)

        from datetime import date
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)

        lease = Lease.objects.create(
            shop=shop,
            tenant=tenant,
            start_date=start_date,
            end_date=end_date,
            monthly_rent=monthly_rent,
            deposit=deposit or 0,
            status='active',
        )
        shop.status = 'occupied'
        shop.save(update_fields=['status'])

        today = timezone.now().date()

        def _make_invoice(inv_type, amount, description, paid=False):
            inv = Invoice.objects.create(
                invoice_number=_generate_invoice_number(),
                tenant=tenant,
                shop=shop,
                invoice_type=inv_type,
                amount=amount,
                issue_date=today,
                due_date=today,
                status='paid' if paid else 'pending',
                paid_date=today if paid else None,
                description=description,
            )
            if paid:
                Payment.objects.create(
                    invoice=inv,
                    amount=amount,
                    payment_date=today,
                    method=payment_method,
                    notes="Paiement initial à la signature du bail",
                )
            return inv

        deposit_val = float(deposit) if deposit else 0
        rent_val = float(monthly_rent)

        if initial_payment in ('deposit', 'both') and deposit_val:
            _make_invoice('charges', deposit_val, f"Caution - Bail {shop}", paid=True)
        elif deposit_val:
            _make_invoice('charges', deposit_val, f"Caution - Bail {shop}", paid=False)

        if initial_payment in ('rent', 'both'):
            _make_invoice('rent', rent_val, f"Premier loyer - {start_date.strftime('%B %Y')}", paid=True)
        else:
            _make_invoice('rent', rent_val, f"Loyer - {start_date.strftime('%B %Y')}", paid=False)

        messages.success(request, f"✅ Bail pour la boutique {shop.shop_number} ({shop.name}) ajouté avec succès à {tenant.full_name}. Factures générées.")
        return redirect('tenant_detail', pk=pk)


class ShopImportTemplateView(LoginRequiredMixin, View):
    """Generates a downloadable Excel template for importing shops and tenants."""
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modèle Importation"
        
        # Header style
        header_fill = openpyxl.styles.PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        headers = [
            "Étage / Niveau",
            "Boutique - Numéro",
            "Boutique - Nom",
            "Boutique - Catégorie",
            "Boutique - Surface (m²)",
            "Boutique - Description",
            "Locataire - Prénom",
            "Locataire - Nom",
            "Locataire - Téléphone",
            "Locataire - Email",
            "Locataire - Raison Sociale",
            "Locataire - N° Pièce d'identité",
            "Bail - Date de début (AAAA-MM-JJ)",
            "Bail - Date de fin (AAAA-MM-JJ)",
            "Bail - Loyer Mensuel (FCFA)",
            "Bail - Caution (FCFA)"
        ]
        
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        # Add a sample row to guide the user
        sample_row = [
            "Rez-de-chaussée",  # Floor
            "A-101",           # Shop Number
            "Boutique Orange",   # Shop Name
            "Télécommunications",# Shop Category
            "25.5",             # Surface
            "Boutique de vente de téléphones et forfaits", # Description
            "Mamadou",          # Tenant First Name
            "Diop",             # Tenant Last Name
            "771234567",        # Tenant Phone
            "mamadou.diop@email.com", # Tenant Email
            "Diop Telecom SARL", # Company Name
            "1234567890123",    # ID number
            "2026-01-01",       # Lease Start
            "2026-12-31",       # Lease End
            "150000",           # Rent
            "300000"            # Deposit
        ]
        ws.append(sample_row)
        
        # Format columns width
        col_widths = [20, 20, 20, 25, 20, 30, 20, 20, 18, 25, 22, 22, 28, 28, 25, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="modele_import_boutiques.xlsx"'
        wb.save(response)
        return response


class ShopImportView(LoginRequiredMixin, View):
    """Handles parsing and validation of uploaded Excel/CSV file to register shops and baux."""
    
    def get(self, request):
        active_mall = request.active_mall
        malls = Mall.objects.all()
        return render(request, 'tenants/shop_import.html', {
            'malls': malls,
            'active_menu': 'shops',
        })
        
    def post(self, request):
        import io
        import csv
        import openpyxl
        from decimal import Decimal
        from datetime import date
        
        active_mall = request.active_mall
        
        # Determine mall
        if not active_mall:
            mall_id = request.POST.get('mall')
            if not mall_id:
                messages.error(request, "Veuillez sélectionner un centre commercial de destination.")
                return redirect('shop_import')
            try:
                mall = Mall.objects.get(pk=mall_id)
            except Mall.DoesNotExist:
                messages.error(request, "Centre commercial sélectionné introuvable.")
                return redirect('shop_import')
        else:
            mall = active_mall
            
        # Options
        overwrite = request.POST.get('overwrite') == 'true'
        initial_invoices = request.POST.get('initial_invoices', 'none')
        
        # File checks
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "Veuillez charger un fichier Excel (.xlsx) ou CSV.")
            return redirect('shop_import')
            
        file_name = uploaded_file.name.lower()
        rows = []
        errors = [] # list of tuples: (row_num, field, value, error_msg)
        
        # 1. READ FILE
        try:
            if file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                # Openpyxl is 1-indexed for rows and cols. Header is row 1, data starts at row 2.
                for r_idx in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 17)]
                    # If the entire row is empty, skip it
                    if not any(v is not None and str(v).strip() != '' for v in row_vals):
                        continue
                    rows.append((r_idx, row_vals))
            elif file_name.endswith('.csv'):
                # Handle CSV
                text_file = io.TextIOWrapper(uploaded_file, encoding='utf-8-sig')
                # Check dialect/delimiter (defaulting to semicolon or comma)
                sample = text_file.read(2048)
                text_file.seek(0)
                delimiter = ';' if ';' in sample else ','
                reader = csv.reader(text_file, delimiter=delimiter)
                header = next(reader, None) # Skip header
                r_idx = 1
                for row_vals in reader:
                    r_idx += 1
                    # Pad to 16 columns if needed
                    row_vals += [None] * (16 - len(row_vals))
                    # Convert empty strings to None
                    row_vals = [v if str(v).strip() != '' else None for v in row_vals]
                    if not any(v is not None for v in row_vals):
                        continue
                    rows.append((r_idx, row_vals))
            else:
                messages.error(request, "Format de fichier non supporté. Veuillez charger un fichier .xlsx ou .csv.")
                return redirect('shop_import')
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier : {str(e)}")
            return redirect('shop_import')
            
        if not rows:
            messages.warning(request, "Le fichier chargé est vide ou ne contient aucune ligne de données.")
            return redirect('shop_import')

        # 2. VALIDATION
        def clean_decimal(val):
            if val is None:
                return None
            val_str = str(val).replace(' ', '').replace('F', '').replace('f', '').replace('\xa0', '').replace(',', '.').replace('\'', '').strip()
            # Strip all non-numeric characters for rents and deposits
            return "".join(c for c in val_str if c.isdigit())
            
        def clean_surface(val):
            if val is None:
                return None
            val_str = str(val).replace(' ', '').replace(',', '.').replace('\'', '').strip()
            # If there's a dot, keep it, but only one
            parts = val_str.split('.')
            if len(parts) > 2:
                val_str = "".join(parts[:-1]) + "." + parts[-1]
            try:
                return Decimal(val_str)
            except:
                return None

        def parse_date_val(val):
            if isinstance(val, (date, timezone.datetime)):
                return val.date() if isinstance(val, timezone.datetime) else val
            if not val:
                return None
            val_str = str(val).strip()
            from django.utils.dateparse import parse_date
            d = parse_date(val_str)
            if d:
                return d
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y', '%d-%m-%y'):
                try:
                    return timezone.datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            return None

        def resolve_category(label):
            if not label:
                return 'other'
            label_clean = str(label).strip().lower()
            
            # Direct keys
            keys = [c[0] for c in Shop.CATEGORY_CHOICES]
            if label_clean in keys:
                return label_clean
                
            # Standard mapping
            mapping = {
                'vêtements & mode': 'clothing',
                'vêtements': 'clothing',
                'mode': 'clothing',
                'vetements': 'clothing',
                'restauration': 'food',
                'restaurant': 'food',
                'nourriture': 'food',
                'électronique': 'electronics',
                'electronique': 'electronics',
                'bijouterie': 'jewelry',
                'bijoux': 'jewelry',
                'beauté & cosmétiques': 'beauty',
                'beauté': 'beauty',
                'beaute': 'beauty',
                'cosmétiques': 'beauty',
                'cosmetiques': 'beauty',
                'sport & loisirs': 'sports',
                'sport': 'sports',
                'loisirs': 'sports',
                'maison & décoration': 'home',
                'maison': 'home',
                'décoration': 'home',
                'decoration': 'home',
                'services': 'services',
                'service': 'services',
                'supermarché': 'supermarket',
                'supermarche': 'supermarket',
                'pharmacie': 'pharmacy',
                'banque & finance': 'bank',
                'banque': 'bank',
                'finance': 'bank',
                'télécommunications': 'telecom',
                'telecommunications': 'telecom',
                'telecom': 'telecom',
                'autre': 'other',
            }
            return mapping.get(label_clean, 'other')

        # Keep track of shop numbers in the uploaded file to detect duplicates in the file itself
        seen_shop_numbers = set()

        for r_idx, row in rows:
            floor_name = row[0]
            shop_number = row[1]
            shop_name = row[2]
            shop_category = row[3]
            shop_surface = row[4]
            tenant_first_name = row[6]
            tenant_last_name = row[7]
            tenant_phone = row[8]
            lease_start_raw = row[12]
            lease_end_raw = row[13]
            lease_rent_raw = row[14]
            
            # Floor check
            if not floor_name:
                errors.append((r_idx, "Étage / Niveau", floor_name, "L'étage est obligatoire."))
                
            # Shop number check
            if not shop_number:
                errors.append((r_idx, "Boutique - Numéro", shop_number, "Le numéro de boutique est obligatoire."))
            else:
                shop_num_str = str(shop_number).strip()
                if shop_num_str in seen_shop_numbers:
                    errors.append((r_idx, "Boutique - Numéro", shop_number, f"Le numéro de boutique '{shop_num_str}' apparaît plusieurs fois dans le fichier."))
                else:
                    seen_shop_numbers.add(shop_num_str)
                    
                # DB check if not overwrite
                if not overwrite:
                    if Shop.objects.filter(shop_number=shop_num_str).exists():
                        errors.append((r_idx, "Boutique - Numéro", shop_number, "Une boutique avec ce numéro existe déjà dans le système (activez l'option d'écrasement pour la modifier)."))
            
            # Shop name check
            if not shop_name:
                errors.append((r_idx, "Boutique - Nom", shop_name, "Le nom de la boutique est obligatoire."))
                
            # Shop surface check
            if shop_surface is not None:
                surface_val = clean_surface(shop_surface)
                if surface_val is None or surface_val <= 0:
                    errors.append((r_idx, "Boutique - Surface", shop_surface, "La surface doit être un nombre décimal positif."))
            else:
                errors.append((r_idx, "Boutique - Surface", shop_surface, "La surface est obligatoire."))
                
            # Tenant occupancy check: if ANY tenant/lease field is filled, we validate lease info
            has_tenant_data = any([tenant_first_name, tenant_last_name, tenant_phone, lease_start_raw, lease_end_raw, lease_rent_raw])
            
            if has_tenant_data:
                if not tenant_first_name:
                    errors.append((r_idx, "Locataire - Prénom", tenant_first_name, "Le prénom du locataire est obligatoire car la boutique est occupée."))
                if not tenant_last_name:
                    errors.append((r_idx, "Locataire - Nom", tenant_last_name, "Le nom du locataire est obligatoire car la boutique est occupée."))
                if not tenant_phone:
                    errors.append((r_idx, "Locataire - Téléphone", tenant_phone, "Le téléphone du locataire est obligatoire car la boutique est occupée."))
                
                # Lease Dates
                start_date = parse_date_val(lease_start_raw)
                end_date = parse_date_val(lease_end_raw)
                
                if not lease_start_raw:
                    errors.append((r_idx, "Bail - Date de début", lease_start_raw, "La date de début de bail est obligatoire."))
                elif not start_date:
                    errors.append((r_idx, "Bail - Date de début", lease_start_raw, "Format de date de début invalide (utilisez AAAA-MM-JJ ou JJ/MM/AAAA)."))
                    
                if not lease_end_raw:
                    errors.append((r_idx, "Bail - Date de fin", lease_end_raw, "La date de fin de bail est obligatoire."))
                elif not end_date:
                    errors.append((r_idx, "Bail - Date de fin", lease_end_raw, "Format de date de fin invalide (utilisez AAAA-MM-JJ ou JJ/MM/AAAA)."))
                    
                if start_date and end_date and end_date <= start_date:
                    errors.append((r_idx, "Bail - Dates", f"{lease_start_raw} à {lease_end_raw}", "La date de fin de bail doit être après la date de début."))
                    
                # Rent
                if not lease_rent_raw:
                    errors.append((r_idx, "Bail - Loyer Mensuel", lease_rent_raw, "Le loyer mensuel est obligatoire."))
                else:
                    rent_clean = clean_decimal(lease_rent_raw)
                    try:
                        rent_val = int(rent_clean) if rent_clean else 0
                        if rent_val <= 0:
                            errors.append((r_idx, "Bail - Loyer Mensuel", lease_rent_raw, "Le loyer doit être un montant positif."))
                    except ValueError:
                        errors.append((r_idx, "Bail - Loyer Mensuel", lease_rent_raw, "Le loyer doit être un nombre entier valide."))
                        
                # Deposit (Caution)
                lease_deposit_raw = row[15]
                if lease_deposit_raw:
                    deposit_clean = clean_decimal(lease_deposit_raw)
                    try:
                        dep_val = int(deposit_clean) if deposit_clean else 0
                        if dep_val < 0:
                            errors.append((r_idx, "Bail - Caution", lease_deposit_raw, "La caution ne peut pas être négative."))
                    except ValueError:
                        errors.append((r_idx, "Bail - Caution", lease_deposit_raw, "La caution doit être un nombre entier valide."))
                        
        # 3. SAVE IF NO ERRORS
        if errors:
            malls = Mall.objects.all()
            return render(request, 'tenants/shop_import.html', {
                'errors': errors,
                'malls': malls,
                'active_menu': 'shops',
            })
            
        # Execute save
        created_shops = 0
        updated_shops = 0
        created_leases = 0
        
        try:
            with transaction.atomic():
                for r_idx, row in rows:
                    floor_name = str(row[0]).strip()
                    shop_number = str(row[1]).strip()
                    shop_name = str(row[2]).strip()
                    shop_category = row[3]
                    shop_surface = clean_surface(row[4])
                    shop_description = str(row[5]).strip() if row[5] else ""
                    
                    tenant_first_name = str(row[6]).strip() if row[6] else None
                    tenant_last_name = str(row[7]).strip() if row[7] else None
                    tenant_phone = str(row[8]).strip() if row[8] else None
                    tenant_email = str(row[9]).strip() if row[9] else ""
                    tenant_company = str(row[10]).strip() if row[10] else ""
                    tenant_id_number = str(row[11]).strip() if row[11] else ""
                    
                    lease_start = parse_date_val(row[12])
                    lease_end = parse_date_val(row[13])
                    
                    lease_rent_clean = clean_decimal(row[14])
                    lease_rent = int(lease_rent_clean) if lease_rent_clean else 0
                    
                    lease_deposit_clean = clean_decimal(row[15]) if row[15] else None
                    lease_deposit = int(lease_deposit_clean) if lease_deposit_clean else 0
                    
                    # 3.1 Get or Create Floor
                    floor, _ = Floor.objects.get_or_create(
                        mall=mall,
                        name=floor_name,
                        defaults={'level': 0}
                    )
                    
                    # 3.2 Get or Create Shop
                    category_code = resolve_category(shop_category)
                    status = 'occupied' if tenant_first_name else 'available'
                    
                    shop_obj = Shop.objects.filter(shop_number=shop_number).first()
                    if shop_obj:
                        # Update existing
                        shop_obj.mall = mall
                        shop_obj.name = shop_name
                        shop_obj.category = category_code
                        shop_obj.floor = floor
                        shop_obj.surface = shop_surface
                        shop_obj.status = status
                        shop_obj.description = shop_description
                        shop_obj.save()
                        updated_shops += 1
                    else:
                        # Create new
                        shop_obj = Shop.objects.create(
                            mall=mall,
                            shop_number=shop_number,
                            name=shop_name,
                            category=category_code,
                            floor=floor,
                            surface=shop_surface,
                            status=status,
                            description=shop_description
                        )
                        created_shops += 1
                        
                    # 3.3 Create Tenant and Lease if occupied
                    if tenant_first_name:
                        # Check if tenant exists by phone
                        tenant_obj = Tenant.objects.filter(phone=tenant_phone).first()
                        if not tenant_obj:
                            tenant_obj = Tenant.objects.create(
                                first_name=tenant_first_name,
                                last_name=tenant_last_name,
                                phone=tenant_phone,
                                email=tenant_email,
                                company_name=tenant_company,
                                id_number=tenant_id_number,
                            )
                        else:
                            # Update tenant email / company if they were empty
                            updated_fields = []
                            if not tenant_obj.email and tenant_email:
                                tenant_obj.email = tenant_email
                                updated_fields.append('email')
                            if not tenant_obj.company_name and tenant_company:
                                tenant_obj.company_name = tenant_company
                                updated_fields.append('company_name')
                            if not tenant_obj.id_number and tenant_id_number:
                                tenant_obj.id_number = tenant_id_number
                                updated_fields.append('id_number')
                            if updated_fields:
                                tenant_obj.save(update_fields=updated_fields)
                                
                        # Create Lease (Terminate existing active leases on this shop just in case)
                        Lease.objects.filter(shop=shop_obj, status='active').update(status='terminated')
                        
                        lease_obj = Lease.objects.create(
                            shop=shop_obj,
                            tenant=tenant_obj,
                            start_date=lease_start,
                            end_date=lease_end,
                            monthly_rent=lease_rent,
                            deposit=lease_deposit,
                            status='active',
                        )
                        created_leases += 1
                        
                        # 3.4 Generate initial invoices if requested
                        if initial_invoices != 'none':
                            is_paid = initial_invoices == 'paid'
                            today_date = timezone.now().date()
                            
                            def _make_invoice(inv_type, amount, desc):
                                inv = Invoice.objects.create(
                                    invoice_number=_generate_invoice_number(),
                                    tenant=tenant_obj,
                                    shop=shop_obj,
                                    invoice_type=inv_type,
                                    amount=amount,
                                    issue_date=today_date,
                                    due_date=today_date,
                                    status='paid' if is_paid else 'pending',
                                    paid_date=today_date if is_paid else None,
                                    description=desc,
                                )
                                if is_paid:
                                    Payment.objects.create(
                                        invoice=inv,
                                        amount=amount,
                                        payment_date=today_date,
                                        method='cash',
                                        notes="Paiement initial enregistré lors de l'importation de masse",
                                    )
                                    
                            if lease_deposit > 0:
                                _make_invoice('charges', lease_deposit, f"Caution - Bail {shop_obj.shop_number}")
                            if lease_rent > 0:
                                _make_invoice('rent', lease_rent, f"Premier loyer - {lease_start.strftime('%B %Y')}")
                                
            messages.success(request, f"✅ Importation réussie ! {created_shops} boutique(s) créée(s), {updated_shops} modifiée(s) et {created_leases} bail(baux) actif(s) enregistré(s).")
            return redirect('shop_list')
        except Exception as e:
            messages.error(request, f"Une erreur imprévue est survenue lors de l'enregistrement en base de données : {str(e)}")
            return redirect('shop_import')

